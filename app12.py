import os
import shutil
import zipfile
from lxml import etree
from openpyxl.utils import coordinate_to_tuple, column_index_from_string, get_column_letter
import re

NS = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

def load_shared_strings(temp_dir):
    """Load shared strings table"""
    shared_strings_path = os.path.join(temp_dir, "xl", "sharedStrings.xml")
    shared_strings = []
    
    if os.path.exists(shared_strings_path):
        try:
            shared_strings_tree = etree.parse(shared_strings_path)
            sst_ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            
            for si in shared_strings_tree.xpath("//ns:si", namespaces=sst_ns):
                t_elem = si.find("ns:t", namespaces=sst_ns)
                if t_elem is not None and t_elem.text:
                    shared_strings.append(t_elem.text)
                else:
                    r_elems = si.findall("ns:r", namespaces=sst_ns)
                    text_parts = []
                    for r_elem in r_elems:
                        t_elem = r_elem.find("ns:t", namespaces=sst_ns)
                        if t_elem is not None and t_elem.text:
                            text_parts.append(t_elem.text)
                    shared_strings.append("".join(text_parts))
            
            print(f"Loaded {len(shared_strings)} shared strings")
        except Exception as e:
            print(f"Error loading shared strings: {e}")
    
    return shared_strings

def get_cell_value_with_shared_strings(cell, shared_strings):
    """Extract cell value from XML element including shared strings"""
    if cell is None:
        return None
        
    cell_type = cell.get("t")
    
    if cell_type == "inlineStr":
        is_elem = cell.find("ns:is", namespaces=NS)
        if is_elem is not None:
            t_elem = is_elem.find("ns:t", namespaces=NS)
            if t_elem is not None and t_elem.text:
                return t_elem.text.strip()
    
    elif cell_type == "s":
        v_elem = cell.find("ns:v", namespaces=NS)
        if v_elem is not None and v_elem.text:
            try:
                string_index = int(v_elem.text)
                if 0 <= string_index < len(shared_strings):
                    return shared_strings[string_index].strip()
            except (ValueError, IndexError):
                pass
    
    elif cell_type == "str" or cell_type is None or cell_type == "":
        v_elem = cell.find("ns:v", namespaces=NS)
        if v_elem is not None and v_elem.text:
            return v_elem.text.strip()
    
    f_elem = cell.find("ns:f", namespaces=NS)
    if f_elem is not None:
        v_elem = cell.find("ns:v", namespaces=NS)
        if v_elem is not None and v_elem.text:
            return v_elem.text.strip()
    
    return None

def find_all_cells_with_content(sheet_tree, shared_strings):
    """Find all cells with content for debugging"""
    sheet_data = sheet_tree.find(".//ns:sheetData", namespaces=NS)
    if sheet_data is None:
        return {}
    
    all_cells = {}
    
    for row in sheet_data.findall("ns:row", namespaces=NS):
        row_num = int(row.get("r", "0"))
        for cell in row.findall("ns:c", namespaces=NS):
            cell_ref = cell.get("r")
            cell_value = get_cell_value_with_shared_strings(cell, shared_strings)
            
            if cell_value and cell_value.strip():
                all_cells[cell_ref] = cell_value.strip()
    
    return all_cells

def find_column_by_header_flexible(sheet_tree, header_names, shared_strings):
    """Find column by header name with flexible matching - prioritize exact matches"""
    sheet_data = sheet_tree.find(".//ns:sheetData", namespaces=NS)
    if sheet_data is None:
        return None, None
    
    merged_ranges, cell_to_merge_map = parse_merged_cells(sheet_tree)
    
    if isinstance(header_names, str):
        header_names = [header_names]
    
    for row in sheet_data.findall("ns:row", namespaces=NS):
        row_num = int(row.get("r", "0"))
        for cell in row.findall("ns:c", namespaces=NS):
            cell_ref = cell.get("r")
            cell_value = get_cell_value_with_shared_strings(cell, shared_strings)
            
            if cell_value:
                cell_value_clean = cell_value.strip().lower()
                
                for header_name in header_names:
                    header_clean = header_name.lower()
                    
                    if cell_value_clean == header_clean:
                        col_letter = ''.join(filter(str.isalpha, cell_ref))
                        print(f"Found '{header_name}' (exact match) at {cell_ref} (Column: {col_letter}, Row: {row_num})")
                        return col_letter, row_num
    
    print(f"No exact match found, searching for partial matches...")
    for row in sheet_data.findall("ns:row", namespaces=NS):
        row_num = int(row.get("r", "0"))
        for cell in row.findall("ns:c", namespaces=NS):
            cell_ref = cell.get("r")
            cell_value = get_cell_value_with_shared_strings(cell, shared_strings)
            
            if cell_value:
                cell_value_clean = cell_value.strip().lower()
                
                for header_name in header_names:
                    header_clean = header_name.lower()
                    
                    if (header_clean in cell_value_clean and len(header_clean) > 3) or \
                       (cell_value_clean in header_clean and len(cell_value_clean) > 3):
                        col_letter = ''.join(filter(str.isalpha, cell_ref))
                        print(f"Found '{header_name}' (partial match: '{cell_value}') at {cell_ref} (Column: {col_letter}, Row: {row_num})")
                        return col_letter, row_num
    
    return None, None

def parse_merged_cells(sheet_tree):
    merged_ranges = []
    cell_to_merge_map = {}
    
    merge_cells_elem = sheet_tree.find(".//ns:mergeCells", namespaces=NS)
    if merge_cells_elem is not None:
        for merge_cell in merge_cells_elem.findall("ns:mergeCell", namespaces=NS):
            ref = merge_cell.get("ref")
            if ref:
                merged_ranges.append(ref)
                
                if ":" in ref:
                    start_cell, end_cell = ref.split(":")
                    start_col, start_row = coordinate_to_tuple(start_cell)
                    end_col, end_row = coordinate_to_tuple(end_cell)
                    
                    for row in range(start_row, end_row + 1):
                        for col in range(start_col, end_col + 1):
                            cell_ref = f"{get_column_letter(col)}{row}"
                            cell_to_merge_map[cell_ref] = start_cell
                else:
                    cell_to_merge_map[ref] = ref
    
    return merged_ranges, cell_to_merge_map

def get_column_values(sheet_tree, col_letter, start_row, shared_strings):
    """Get all non-empty values from a column starting from a specific row"""
    sheet_data = sheet_tree.find(".//ns:sheetData", namespaces=NS)
    if sheet_data is None:
        return []
    
    column_values = []
    col_index = column_index_from_string(col_letter)
    
    for row in sheet_data.findall("ns:row", namespaces=NS):
        row_num = int(row.get("r", "0"))
        
        if row_num <= start_row:
            continue
            
        for cell in row.findall("ns:c", namespaces=NS):
            cell_ref = cell.get("r")
            cell_col_letter = ''.join(filter(str.isalpha, cell_ref))
            cell_col_index = column_index_from_string(cell_col_letter)
            
            if cell_col_index == col_index:
                cell_value = get_cell_value_with_shared_strings(cell, shared_strings)
                if cell_value and cell_value.strip():
                    column_values.append({
                        'value': cell_value.strip(),
                        'row': row_num,
                        'cell_ref': cell_ref
                    })
                break
    
    column_values.sort(key=lambda x: x['row'])
    return column_values

def create_mapping_for_analysis_column(items_values, analysis_col, keyword_map):
    """Create mapping for analysis column based on items values"""
    cell_value_mapping = {}
    
    for item_info in items_values:
        item_value = item_info['value']
        item_row = item_info['row']
        
        mapped_value = keyword_map.get(item_value)
        
        if not mapped_value:
            for key, value in keyword_map.items():
                if key.lower() in item_value.lower() or item_value.lower() in key.lower():
                    mapped_value = value
                    print(f"  Partial match found: '{item_value}' -> '{key}' -> '{mapped_value}'")
                    break
        
        if mapped_value:
            analysis_cell_ref = f"{analysis_col}{item_row}"
            cell_value_mapping[analysis_cell_ref] = mapped_value
            print(f"  Mapping: {item_value} -> {mapped_value} at {analysis_cell_ref}")
        else:
            print(f"  Warning: No mapping found for '{item_value}' in keyword_map")
    
    return cell_value_mapping

def update_analysis_cells(source_path, destination_folder, keyword_map):
    os.makedirs(destination_folder, exist_ok=True)
    file_name = os.path.basename(source_path)
    dest_path = os.path.join(destination_folder, file_name)
    shutil.copy2(source_path, dest_path)
    print(f"Copied file to: {dest_path}")

    temp_dir = os.path.join(destination_folder, "temp_unzip")
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
    os.makedirs(temp_dir)

    try:
        with zipfile.ZipFile(dest_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        print("Excel file extracted successfully")

        shared_strings = load_shared_strings(temp_dir)

        workbook_xml = os.path.join(temp_dir, "xl", "workbook.xml")
        if not os.path.exists(workbook_xml):
            raise FileNotFoundError("workbook.xml not found")
            
        tree = etree.parse(workbook_xml)
        sheet_id = None
        
        for sheet in tree.xpath("//ns:sheets/ns:sheet", namespaces=NS):
            if sheet.get("name") == "07.Analysis":
                sheet_id = sheet.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                print(f"Found sheet '07.Analysis' with ID: {sheet_id}")
                break
                
        if not sheet_id:
            available_sheets = []
            for sheet in tree.xpath("//ns:sheets/ns:sheet", namespaces=NS):
                available_sheets.append(sheet.get("name"))
            print(f"Available sheets: {available_sheets}")
            raise ValueError("Sheet '07.Analysis' not found in the workbook")

        rels_path = os.path.join(temp_dir, "xl", "_rels", "workbook.xml.rels")
        if not os.path.exists(rels_path):
            raise FileNotFoundError("workbook.xml.rels not found")
            
        rels_tree = etree.parse(rels_path)
        rels_ns = {"ns": "http://schemas.openxmlformats.org/package/2006/relationships"}
        sheet_file = None
        
        for rel in rels_tree.xpath("//ns:Relationship", namespaces=rels_ns):
            if rel.get("Id") == sheet_id:
                sheet_file = rel.get("Target").split("/")[-1]
                print(f"Found worksheet file: {sheet_file}")
                break
                
        if not sheet_file:
            raise ValueError("Cannot find sheet file for '07.Analysis'")

        sheet_xml_path = os.path.join(temp_dir, "xl", "worksheets", sheet_file)
        if not os.path.exists(sheet_xml_path):
            raise FileNotFoundError(f"Worksheet file {sheet_file} not found")
            
        parser = etree.XMLParser(resolve_entities=False)
        sheet_tree = etree.parse(sheet_xml_path, parser)
        
        print(f"Parsed worksheet XML successfully")
        
        sheet_data = sheet_tree.find(".//ns:sheetData", namespaces=NS)
        if sheet_data is None:
            raise ValueError("sheetData element not found in worksheet")

        merged_ranges, cell_to_merge_map = parse_merged_cells(sheet_tree)
        print(f"Found {len(merged_ranges)} total merged cell ranges")

        print("\n=== DEBUG: All cells with content ===")
        all_cells = find_all_cells_with_content(sheet_tree, shared_strings)
        for cell_ref, value in sorted(all_cells.items()):
            print(f"  {cell_ref}: '{value}'")
        print("=== END DEBUG ===\n")

        items_col, items_header_row = find_column_by_header_flexible(
            sheet_tree, 
            ["Items", "Item", "Item Name", "Item Type", "Test Items"], 
            shared_strings
        )
        
        if not items_col:
            raise ValueError("'Items' column not found. Please check the debug output above to see all cell values.")
        
        analysis_col, analysis_header_row = find_column_by_header_flexible(
            sheet_tree, 
            ["Analysis", "Analyse", "Result", "Results", "Status"], 
            shared_strings
        )
        
        if not analysis_col:
            raise ValueError("'Analysis' column not found. Please check the debug output above to see all cell values.")
        
        print(f"Items column: {items_col}, Analysis column: {analysis_col}")
        
        items_values = get_column_values(sheet_tree, items_col, items_header_row, shared_strings)
        print(f"Found {len(items_values)} items in Items column:")
        for item in items_values:
            print(f"  Row {item['row']}: '{item['value']}'")
        
        print(f"\nCreating mappings using keyword_map:")
        cell_value_mapping = create_mapping_for_analysis_column(items_values, analysis_col, keyword_map)
        
        if not cell_value_mapping:
            print("No mappings created. Please check your keyword_map and Items column values.")
            return
        
        print(f"\nStarting to update {len(cell_value_mapping)} cells...")
        
        replaced_count = 0
        created_rows = 0
        created_cells = 0
        
        for cell_ref, val in cell_value_mapping.items():
            row_num = int(''.join(filter(str.isdigit, cell_ref)))
            
            print(f"Processing cell {cell_ref} with value '{val}'")
            
            row_element = None
            for row in sheet_data.findall("ns:row", namespaces=NS):
                if int(row.get("r", "0")) == row_num:
                    row_element = row
                    break
            
            if row_element is None:
                row_element = etree.Element(f"{{{NS['ns']}}}row")
                row_element.set("r", str(row_num))
                
                inserted = False
                for j, existing_row in enumerate(sheet_data.findall("ns:row", namespaces=NS)):
                    existing_row_num = int(existing_row.get("r", "0"))
                    if existing_row_num > row_num:
                        sheet_data.insert(j, row_element)
                        inserted = True
                        break
                
                if not inserted:
                    sheet_data.append(row_element)
                
                created_rows += 1
                print(f"  Created new row {row_num}")
            
            cell = None
            for c in row_element.findall("ns:c", namespaces=NS):
                if c.get("r") == cell_ref:
                    cell = c
                    break
            
            if cell is None:
                cell = etree.Element(f"{{{NS['ns']}}}c")
                cell.set("r", cell_ref)
                
                col_letter_only = ''.join(filter(str.isalpha, cell_ref))
                col_index = column_index_from_string(col_letter_only)
                
                inserted = False
                for j, existing_cell in enumerate(row_element.findall("ns:c", namespaces=NS)):
                    existing_ref = existing_cell.get("r", "A1")
                    existing_col_str = ''.join(filter(str.isalpha, existing_ref))
                    existing_col_index = column_index_from_string(existing_col_str)
                    
                    if existing_col_index > col_index:
                        row_element.insert(j, cell)
                        inserted = True
                        break
                
                if not inserted:
                    row_element.append(cell)
                
                created_cells += 1
                print(f"  Created new cell {cell_ref}")
            
            for child in list(cell):
                cell.remove(child)
            
            cell.set("t", "inlineStr")
            
            is_element = etree.SubElement(cell, f"{{{NS['ns']}}}is")
            t_element = etree.SubElement(is_element, f"{{{NS['ns']}}}t")
            t_element.text = str(val)
            
            replaced_count += 1
            print(f"  Successfully updated {cell_ref} = '{val}'")

        print(f"\nSummary:")
        print(f"  - Updated {replaced_count} cells")
        print(f"  - Created {created_rows} new rows") 
        print(f"  - Created {created_cells} new cells")

        with open(sheet_xml_path, 'wb') as f:
            sheet_tree.write(f, xml_declaration=True, encoding="UTF-8", standalone="yes")
        
        print(f"Worksheet {sheet_file} saved successfully")

        if os.path.exists(dest_path):
            os.remove(dest_path)
        
        with zipfile.ZipFile(dest_path, 'w', zipfile.ZIP_DEFLATED) as zip_out:
            for root_dir, _, files in os.walk(temp_dir):
                for f in files:
                    full_path = os.path.join(root_dir, f)
                    rel_path = os.path.relpath(full_path, temp_dir)
                    rel_path = rel_path.replace(os.path.sep, '/')
                    zip_out.write(full_path, rel_path)
        
        print(f"Excel file repacked successfully: {dest_path}")

    except Exception as e:
        print(f"Error occurred: {str(e)}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            print("Temporary files cleaned up")

def validate_excel_file(file_path):
    try:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            file_list = zip_ref.namelist()
            required_files = [
                'xl/workbook.xml',
                'xl/_rels/workbook.xml.rels',
                '[Content_Types].xml'
            ]
            
            missing_files = []
            for req_file in required_files:
                if req_file not in file_list:
                    missing_files.append(req_file)
            
            if missing_files:
                print(f"Warning: Missing required files: {missing_files}")
                return False
            
        print("Excel file structure validation passed")
        return True
    except Exception as e:
        print(f"Excel file validation failed: {str(e)}")
        return False

def get_user_inputs():
    print("\n=== Excel Analysis Column Updater ===")
    source_excel=r"C:\Users\kumarsaw\Downloads\RJIL 5G Cluster Acceptance Report_KA-HBDD-0002_7_3500.xlsx"
    # source_excel = input("Enter source Excel file path: ").strip()
    if source_excel.startswith('"') and source_excel.endswith('"'):
        source_excel = source_excel[1:-1]
    
    # destination_folder = input("Enter destination folder path: ").strip()
    destination_folder="."
    if destination_folder.startswith('"') and destination_folder.endswith('"'):
        destination_folder = destination_folder[1:-1]
    
    return source_excel, destination_folder

if __name__ == "__main__":
    
    keyword_map = {
        "Tilt": "Tilt Report",
        "GPL": "Pre Post",
        "Tilt Table": "Tilt Report",
        "MRJ": "MRJ is attached", 
        "Swap Check": "Swap report is added"
       
    }
    
    try:
        source_excel, destination_folder = get_user_inputs()
        
        print(f"\nConfiguration:")
        print(f"Source: {source_excel}")
        print(f"Destination: {destination_folder}")
        print(f"Keyword mappings: {keyword_map}")      
        
        
        if not os.path.exists(source_excel):
            print(f"Error: Source file does not exist: {source_excel}")
            exit(1)
        
        if not validate_excel_file(source_excel):
            print("Source file validation failed. Please check the file.")
            exit(1)
        
        update_analysis_cells(source_excel, destination_folder, keyword_map)
        
        output_file = os.path.join(destination_folder, os.path.basename(source_excel))
        if validate_excel_file(output_file):
            print("\n✅ SUCCESS: File processed successfully!")
            print(f"Output file: {output_file}")
        else:
            print("\n❌ WARNING: Output file may have issues")
            
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user.")
        exit(0)
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        exit(1)