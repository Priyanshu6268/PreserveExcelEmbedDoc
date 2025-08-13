import os
import shutil
import zipfile
from lxml import etree
from openpyxl.utils import coordinate_to_tuple, column_index_from_string
import re

NS = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

def parse_merged_cells(sheet_tree):
    merged_ranges = []
    cell_to_merge_map = {}
    
    merge_cells_elem = sheet_tree.find(".//ns:mergeCells", namespaces=NS)
    if merge_cells_elem is not None:
        for merge_cell in merge_cells_elem.findall("ns:mergeCell", namespaces=NS):
            ref = merge_cell.get("ref")
            if ref:
                merged_ranges.append(ref)
                
                if ":" in ref:
                    start_cell, end_cell = ref.split(":")
                    start_col, start_row = coordinate_to_tuple(start_cell)
                    end_col, end_row = coordinate_to_tuple(end_cell)
                    
                    for row in range(start_row, end_row + 1):
                        for col in range(start_col, end_col + 1):
                            from openpyxl.utils import get_column_letter
                            cell_ref = f"{get_column_letter(col)}{row}"
                            cell_to_merge_map[cell_ref] = start_cell
                else:
                    cell_to_merge_map[ref] = ref
    
    return merged_ranges, cell_to_merge_map

def get_merged_ranges_for_target_column(merged_ranges, target_column):
    target_ranges = []
    target_col_idx = column_index_from_string(target_column)
    
    for merge_range in merged_ranges:
        if ":" not in merge_range:
            continue
            
        start_cell, end_cell = merge_range.split(":")
        start_col_str = ''.join(filter(str.isalpha, start_cell))
        start_row_num = int(''.join(filter(str.isdigit, start_cell)))
        
        start_col_idx = column_index_from_string(start_col_str)
        
        if start_col_idx == target_col_idx:
            end_col_str = ''.join(filter(str.isalpha, end_cell))
            end_row_num = int(''.join(filter(str.isdigit, end_cell)))
            
            block_size = end_row_num - start_row_num + 1
            target_ranges.append({
                'range': merge_range,
                'start_cell': start_cell,
                'start_row': start_row_num,
                'end_row': end_row_num,
                'block_size': block_size
            })
    
    target_ranges.sort(key=lambda x: x['start_row'])
    return target_ranges

def map_values_to_merged_cells_fixed(cluster_values, target_ranges, start_row):
    cell_value_mapping = {}
    value_index = 0
    
    print(f"Mapping {len(cluster_values)} values to cells starting from row {start_row}")
    print(f"Found {len(target_ranges)} merged ranges starting with target column")
    
    sorted_ranges = sorted(target_ranges, key=lambda x: x['start_row'])
    
    current_row = start_row
    range_index = 0
    
    while value_index < len(cluster_values):
        if value_index >= len(cluster_values):
            break
            
        current_merge_range = None
        for merge_info in sorted_ranges:
            if merge_info['start_row'] <= current_row <= merge_info['end_row']:
                current_merge_range = merge_info
                break
        
        if current_merge_range:
            if current_row == current_merge_range['start_row']:
                top_left_cell = current_merge_range['start_cell']
                cell_value_mapping[top_left_cell] = cluster_values[value_index]
                print(f"  Merged range {current_merge_range['range']}: {top_left_cell} = '{cluster_values[value_index]}'")
                value_index += 1
            
            current_row = current_merge_range['end_row'] + 1
        else:
            target_col = ''.join(filter(str.isalpha, list(cell_value_mapping.keys())[0] if cell_value_mapping else "AG"))
            if not target_col:
                target_col = "AG"
            cell_ref = f"{target_col}{current_row}"
            cell_value_mapping[cell_ref] = cluster_values[value_index]
            print(f"  Individual cell {cell_ref} = '{cluster_values[value_index]}'")
            value_index += 1
            current_row += 1
    
    print(f"  Total cells to update: {len(cell_value_mapping)}")
    print(f"  Used {value_index} out of {len(cluster_values)} values")
    
    return cell_value_mapping

def replace_existing_cells(source_path, destination_folder, cluster_values, start_cell):
    os.makedirs(destination_folder, exist_ok=True)
    file_name = os.path.basename(source_path)
    dest_path = os.path.join(destination_folder, file_name)
    shutil.copy2(source_path, dest_path)
    print(f"Copied file to: {dest_path}")

    temp_dir = os.path.join(destination_folder, "temp_unzip")
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
    os.makedirs(temp_dir)

    try:
        with zipfile.ZipFile(dest_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        print("Excel file extracted successfully")

        workbook_xml = os.path.join(temp_dir, "xl", "workbook.xml")
        if not os.path.exists(workbook_xml):
            raise FileNotFoundError("workbook.xml not found")
            
        tree = etree.parse(workbook_xml)
        sheet_id = None
        
        for sheet in tree.xpath("//ns:sheets/ns:sheet", namespaces=NS):
            if sheet.get("name") == "07.Analysis":
                sheet_id = sheet.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                print(f"Found sheet '07.Analysis' with ID: {sheet_id}")
                break
                
        if not sheet_id:
            available_sheets = []
            for sheet in tree.xpath("//ns:sheets/ns:sheet", namespaces=NS):
                available_sheets.append(sheet.get("name"))
            print(f"Available sheets: {available_sheets}")
            raise ValueError("Sheet '07.Analysis' not found in the workbook")

        rels_path = os.path.join(temp_dir, "xl", "_rels", "workbook.xml.rels")
        if not os.path.exists(rels_path):
            raise FileNotFoundError("workbook.xml.rels not found")
            
        rels_tree = etree.parse(rels_path)
        rels_ns = {"ns": "http://schemas.openxmlformats.org/package/2006/relationships"}
        sheet_file = None
        
        for rel in rels_tree.xpath("//ns:Relationship", namespaces=rels_ns):
            if rel.get("Id") == sheet_id:
                sheet_file = rel.get("Target").split("/")[-1]
                print(f"Found worksheet file: {sheet_file}")
                break
                
        if not sheet_file:
            raise ValueError("Cannot find sheet file for '07.Analysis'")

        sheet_xml_path = os.path.join(temp_dir, "xl", "worksheets", sheet_file)
        if not os.path.exists(sheet_xml_path):
            raise FileNotFoundError(f"Worksheet file {sheet_file} not found")
            
        parser = etree.XMLParser(resolve_entities=False)
        sheet_tree = etree.parse(sheet_xml_path, parser)
        
        print(f"Parsed worksheet XML successfully")
        
        sheet_data = sheet_tree.find(".//ns:sheetData", namespaces=NS)
        if sheet_data is None:
            raise ValueError("sheetData element not found in worksheet")

        merged_ranges, cell_to_merge_map = parse_merged_cells(sheet_tree)
        print(f"Found {len(merged_ranges)} total merged cell ranges")

        col_letter = ''.join(filter(str.isalpha, start_cell))
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        
        target_ranges = get_merged_ranges_for_target_column(merged_ranges, col_letter)
        
        print(f"\nMerged ranges starting with column {col_letter}:")
        for i, merge_info in enumerate(target_ranges):
            print(f"  {i+1}. {merge_info['range']} (rows {merge_info['start_row']}-{merge_info['end_row']}, size: {merge_info['block_size']})")
        
        cell_value_mapping = map_values_to_merged_cells_fixed(cluster_values, target_ranges, start_row)
        
        print(f"\nStarting to update cells...")
        
        replaced_count = 0
        created_rows = 0
        created_cells = 0
        
        for cell_ref, val in cell_value_mapping.items():
            row_num = int(''.join(filter(str.isdigit, cell_ref)))
            
            print(f"Processing cell {cell_ref} with value '{val}'")
            
            row_element = None
            for row in sheet_data.findall("ns:row", namespaces=NS):
                if int(row.get("r", "0")) == row_num:
                    row_element = row
                    break
            
            if row_element is None:
                row_element = etree.Element(f"{{{NS['ns']}}}row")
                row_element.set("r", str(row_num))
                
                inserted = False
                for j, existing_row in enumerate(sheet_data.findall("ns:row", namespaces=NS)):
                    existing_row_num = int(existing_row.get("r", "0"))
                    if existing_row_num > row_num:
                        sheet_data.insert(j, row_element)
                        inserted = True
                        break
                
                if not inserted:
                    sheet_data.append(row_element)
                
                created_rows += 1
                print(f"  Created new row {row_num}")
            
            cell = None
            for c in row_element.findall("ns:c", namespaces=NS):
                if c.get("r") == cell_ref:
                    cell = c
                    break
            
            if cell is None:
                cell = etree.Element(f"{{{NS['ns']}}}c")
                cell.set("r", cell_ref)
                
                col_letter_only = ''.join(filter(str.isalpha, cell_ref))
                col_index = column_index_from_string(col_letter_only)
                
                inserted = False
                for j, existing_cell in enumerate(row_element.findall("ns:c", namespaces=NS)):
                    existing_ref = existing_cell.get("r", "A1")
                    existing_col_str = ''.join(filter(str.isalpha, existing_ref))
                    existing_col_index = column_index_from_string(existing_col_str)
                    
                    if existing_col_index > col_index:
                        row_element.insert(j, cell)
                        inserted = True
                        break
                
                if not inserted:
                    row_element.append(cell)
                
                created_cells += 1
                print(f"  Created new cell {cell_ref}")
            
            for child in list(cell):
                cell.remove(child)
            
            cell.set("t", "inlineStr")
            
            is_element = etree.SubElement(cell, f"{{{NS['ns']}}}is")
            t_element = etree.SubElement(is_element, f"{{{NS['ns']}}}t")
            t_element.text = str(val)
            
            replaced_count += 1
            print(f"  Successfully updated {cell_ref} = '{val}'")

        print(f"\nSummary:")
        print(f"  - Updated {replaced_count} cells")
        print(f"  - Created {created_rows} new rows") 
        print(f"  - Created {created_cells} new cells")
        print(f"  - Processed {len(target_ranges)} merged cell ranges")

        with open(sheet_xml_path, 'wb') as f:
            sheet_tree.write(f, xml_declaration=True, encoding="UTF-8", standalone="yes")
        
        print(f"Worksheet {sheet_file} saved successfully")

        if os.path.exists(dest_path):
            os.remove(dest_path)
        
        with zipfile.ZipFile(dest_path, 'w', zipfile.ZIP_DEFLATED) as zip_out:
            for root_dir, _, files in os.walk(temp_dir):
                for f in files:
                    full_path = os.path.join(root_dir, f)
                    rel_path = os.path.relpath(full_path, temp_dir)
                    rel_path = rel_path.replace(os.path.sep, '/')
                    zip_out.write(full_path, rel_path)
        
        print(f"Excel file repacked successfully: {dest_path}")

    except Exception as e:
        print(f"Error occurred: {str(e)}")
        import traceback
        traceback.print_exc()
        raise
    finally:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
            print("Temporary files cleaned up")

def validate_excel_file(file_path):
    try:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            file_list = zip_ref.namelist()
            required_files = [
                'xl/workbook.xml',
                'xl/_rels/workbook.xml.rels',
                '[Content_Types].xml'
            ]
            
            missing_files = []
            for req_file in required_files:
                if req_file not in file_list:
                    missing_files.append(req_file)
            
            if missing_files:
                print(f"Warning: Missing required files: {missing_files}")
                return False
            
        print("Excel file structure validation passed")
        return True
    except Exception as e:
        print(f"Excel file validation failed: {str(e)}")
        return False

def get_user_inputs():
    # print("\n=== Excel Cell Replacement Tool ===")
    
    # source_excel = input("Enter source Excel file path: ").strip()
    # if source_excel.startswith('"') and source_excel.endswith('"'):
    #     source_excel = source_excel[1:-1]
    
    # destination_folder = input("Enter destination folder path: ").strip()
    # if destination_folder.startswith('"') and destination_folder.endswith('"'):
    #     destination_folder = destination_folder[1:-1]

    source_excel = r"C:\Users\kumarsaw\Downloads\RJIL 5G Cluster Acceptance Report_KA-HBDD-0002_7_3500 - Copy.xlsx"
    destination_folder = r"C:\Users\kumarsaw\Documents"

    
    start_cell = input("Enter starting cell (e.g., AG11): ").strip().upper()
    
    print("\nEnter cluster values (one per line). Press Enter twice when done:")
    cluster_values = []
    while True:
        value = input(f"Value {len(cluster_values) + 1}: ").strip()
        if not value:
            if cluster_values:
                break
            else:
                print("Please enter at least one value.")
                continue
        cluster_values.append(value)
    
    return source_excel, destination_folder, start_cell, cluster_values

if __name__ == "__main__":
    try:
        source_excel, destination_folder, start_cell, cluster_values = get_user_inputs()
        
        print(f"\nConfiguration:")
        print(f"Source: {source_excel}")
        print(f"Destination: {destination_folder}")
        print(f"Starting cell: {start_cell}")
        print(f"Values to insert: {len(cluster_values)} items - {cluster_values}")
        
        confirm = input("\nProceed with these settings? (y/n): ").strip().lower()
        if confirm != 'y':
            print("Operation cancelled.")
            exit(0)
        
        if not os.path.exists(source_excel):
            print(f"Error: Source file does not exist: {source_excel}")
            exit(1)
        
        if not validate_excel_file(source_excel):
            print("Source file validation failed. Please check the file.")
            exit(1)
        
        replace_existing_cells(source_excel, destination_folder, cluster_values, start_cell)
        
        output_file = os.path.join(destination_folder, os.path.basename(source_excel))
        if validate_excel_file(output_file):
            print("\n✅ SUCCESS: File processed successfully!")
            print(f"Output file: {output_file}")
        else:
            print("\n❌ WARNING: Output file may have issues")
            
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user.")
        exit(0)
    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        exit(1)